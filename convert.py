# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
import os


def getValue(val):
    
    if not val:
        return ''

    return '{0}'.format(val)


FILENAME = 'd:\\tmp\\rubcov\\файл2.xlsx'
OUTFILE =  'd:\\tmp\\rubcov\\2.csv'
DELIMENTER = '~'
NAMESHEET = 'TDSheet'

wb = load_workbook(FILENAME, read_only = True)
sh = wb[NAMESHEET]

i = 0

with open(OUTFILE, "w", encoding = 'utf-8') as dest:
    for row in sh.iter_rows():

        i += 1
        if i % 500 == 0:
            print('left {0}'.format(i))

        data = [getValue(cell.value) for cell in row]

        newStr = DELIMENTER.join(data)

        try:
            dest.write(u'{0}\n'.format(newStr))
        except UnicodeEncodeError:
            print(newStr)
            exit()

wb.close()

