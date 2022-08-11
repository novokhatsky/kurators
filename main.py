# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from openpyxl.cell import WriteOnlyCell
import os
from copy import copy
from datetime import date
from datetime import datetime
import shutil
import traceback

BASE_DIR = "d:\\work\\in\\"
BASE_OUT = "d:\\work\\out\\"
BACKUP_PATH = "d:\\work\\backup\\"
#BASE_DIR = "d:\\tmp\\rubcov\\pro\\in\\"
#BASE_OUT = "d:\\tmp\\rubcov\\pro\\out\\"
#BACKUP_PATH = "d:\\tmp\\rubcov\\pro\\backup\\"

DIFF_PATH = BASE_OUT + "diff\\"

PPR = BASE_DIR + "пэн-ппр\\ППР.xlsx"
PEN = BASE_DIR + "пэн-ппр\\ПЭН.xlsx"

PPR_OUT = BASE_OUT + "пэн-ппр\\ППР.xlsx"
PEN_OUT = BASE_OUT + "пэн-ппр\\ПЭН.xlsx"

# поля которые переносятся из ПЭН и ППР в файлы кураторов
PPR_INDEX_I = [21, 22, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 74]
PEN_INDEX_I = [21, 22, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 74]

# поля которые переносятся из кураторов в ПЭН и ППР
PPR_INDEX_II = [19, 20, 50, 51, 52, 53, 54, 55, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]
PEN_INDEX_II = [19, 20, 50, 51, 52, 53, 54, 55, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]

CHECK_HEADER = 84

def makeOut(filename):
    el = filename.split("\\")

    return BASE_OUT + el[-1]


class WrongHeader(Exception):
    def __init__(self, org, curr, pos, message = ''):
        self.org = org
        self.curr = curr
        self.pos = pos + 1 
        self.message = "несовпадение названия колонки {2}, образец: '{0}' текущее: '{1}'".format(self.org, self.curr, self.pos)

    def __str__(self):
        return self.message 


class NotUnique(Exception):
    def __init__(self, value, message = ''):
        self.value = value
        self.message = "неуникальное значение '{0}'".format(self.value)

    def __str__(self):
        return self.message 


class Unique():
    def __init__(self):
        self.data = []

    def add(self, value):
        if value is None:
            return

        if value in self.data:
            raise NotUnique(value)
        
        self.data.append(value)


def checkHeader(header):
    if checkHeader.template == []:
        checkHeader.template = header
        return

    for i in range(CHECK_HEADER):
        if checkHeader.template[i] != header[i]:
            raise WrongHeader(checkHeader.template[i], header[i], i) 
            

def ExitWMessage(message):
    print(message)
    input("Для выхода нажмите Enter")
    exit()


def makeDict(filename):
    wb = load_workbook(filename, read_only = True, data_only = True)
    sh = wb.active
    data = {}
    num_str = 0
    unique = Unique()

    # нужно пропустить три строки
    enable_add = False
    need_len = 0

    for row in sh.iter_rows():
        num_str += 1
    
        if enable_add:

            try:
                unique.add(row[0].value)
            except NotUnique as e:
                ExitWMessage("ошибка в файле {0} в строке {2}\n{1}".format(filename, e.message, num_str))

            data[row[0].value] = [cell.value for cell in row]
            curr_len = len(data[row[0].value])

            if curr_len < need_len:
                # выравниваем размер массива
                data[row[0].value].extend(['' for i in range(need_len - curr_len)])

            continue

        if len(row) > 0 and row[0].value == 'Идентификатор':
            need_len = len(row)
            enable_add = True

            try:
                checkHeader([cell.value for cell in row])
            except WrongHeader as e:
                ExitWMessage("ошибка в файле {0}\n{1}".format(filename, e.message))

    wb.close()

    return data


def listFiles(base_dir):
    # формируем список файлов с заданной маской в директории
    spisok_file = []
    for i in os.listdir(base_dir):
        if os.path.isfile(os.path.join(base_dir, i)):
           if i.endswith('.xlsx') or i.endswith('.xlsm'):
                spisok_file.append(os.path.join(base_dir, i))

    return spisok_file


def currDateTime():
    return 'Обновлено: ' + datetime.today().strftime('%d-%m-%Y %H:%M')


def makeBackup():
    fullFileBackup = BACKUP_PATH + date.today().isoformat()

    if os.path.isdir(fullFileBackup):
        print("на текущее число копия уже есть")
        return

    # создаем каталог и переносим все вайлы из in
    shutil.copytree(BASE_DIR, fullFileBackup)

    # удаляем каталог ИН
    shutil.rmtree(BASE_DIR)

    # создаем каталог ИН и копируем туда данные из АУТ
    shutil.copytree(BASE_OUT, BASE_DIR)

    # удаляем каталог АУТ 
    shutil.rmtree(BASE_OUT)

    # создаем АУТ и пэн-ппр
    os.mkdir(BASE_OUT)
    os.mkdir(BASE_OUT + "пэн-ппр")


class fileExcel(object):
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None
        self.sheet = None


    def create(self, sheetname):
        self.workbook = Workbook(write_only = True)
        self.sheet = self.workbook[self.workbook.create_sheet(sheetname).title]


    def append(self, row):
        self.sheet.append(row)


    def save(self):
        self.workbook.save(self.filename)
        self.workbook.close()


class kuratorsDict(object):
    def __init__(self):
        self.dicts = []


    def load(self):
        for fl in listFiles(BASE_OUT):
            print('загрузка {0}'.format(fl))
            self.dicts.append(makeDict(fl))


    def seekkey(self, key):
        for dic in self.dicts:

            if key in dic.keys():

                return dic[key]

        return False


class kuratorsCheck(object):
    def __init__(self, fileIn, fileOut, notFound, dicts):
        print('загрузка ' + fileIn)
        self.wb = load_workbook(fileIn, read_only = True, data_only = True)
        self.sh = self.wb.active

        # создаем книгу для создания нового файла ПЭН или ППР
        self.fileOut = fileExcel(fileOut)
        self.fileOut.create('ПЭН_ППР сокращ')

        # создаем книгу для записи ненайденных ИД
        self.notFoundId = fileExcel(notFound)
        self.notFoundId.create('not found')

        # создание словаря кураторов
        self.dicts = dicts

        self.fileIn = fileIn

    def seekChange(self, indexCells):
        print('поиск')
        i = 0
        unique = Unique()
        enable_seek = False
        need_len = 0

        for row in self.sh.iter_rows():
            # если в cтроке нет элементов, пропускаем
            if len(row) < 1:
                continue

            i += 1
            if i % 500 == 0:
                print('обработано {0}'.format(i))

            key = row[0].value

            if enable_seek:

                try:
                    unique.add(row[0].value)
                except NotUnique as e:
                    ExitWMessage("ошибка в файле {0}\n{1}".format(self.fileIn, e.message))

                found = self.dicts.seekkey(key)

                if found:
                    new_row = [cell.value for cell in row]

                    if len(new_row) < need_len:
                        new_row.extend(['' for i in range(need_len - len(new_row))])

                    for index in indexCells:
                        new_row[index - 1] = found[index - 1]

                    self.fileOut.append(new_row)

                else:
                    self.fileOut.append([cell.value for cell in row])
                    self.notFoundId.append([key])

                continue                

            if key == 'Идентификатор':
                need_len = len(row)
                enable_seek = True

                try:
                    checkHeader([cell.value for cell in row])
                except WrongHeader as e:
                    ExitWMessage("ошибка в файле {0}\n{1}".format(self.fileIn, e.message))

            # копируем шапку со стилями
            new_row = []
            for cell in row:
                new_cell = WriteOnlyCell(self.fileOut.sheet, cell.value)

                if cell.fill:
                    new_cell.fill = copy(cell.fill)

                if cell.font:
                    new_cell.font = copy(cell.font)

                if cell.border:
                    new_cell.border = copy(cell.border)

                if cell.alignment:
                    new_cell.alignment = copy(cell.alignment)

                new_row.append(new_cell)

            self.fileOut.append([cell for cell in new_row])

        print('завершено на {0}'.format(i))

    def save(self):
        print('save')
        del self.sh
        self.wb.close()

        self.notFoundId.save()
        self.fileOut.save()


def updatePprPen():
    print('загрузка файлов кураторов:')
    kurators_dict = kuratorsDict()
    kurators_dict.load()

    if os.path.isfile(PPR):
        ppr = kuratorsCheck(PPR, PPR_OUT, DIFF_PATH + 'no_id_ppr_in_kur.xlsx', kurators_dict)
        ppr.seekChange(PPR_INDEX_II)
        ppr.save()

    if os.path.isfile(PEN):
        pen = kuratorsCheck(PEN, PEN_OUT, DIFF_PATH + 'no_id_pen_in_kur.xlsx', kurators_dict)
        pen.seekChange(PEN_INDEX_II)
        pen.save()


def updateKurators():

    # если файла ППР нет, используем пустой справочник, иначе загружаем справочник из файла
    if os.path.isfile(PPR):
        print('загрузка {0}, используем название столбцов как образец'.format(PPR))
        ppr_dict = makeDict(PPR)
    else:
        ppr_dict = {}


    # если файла ПЕН нет, используем пустой справочник, иначе загружаем справочник из файла
    if os.path.isfile(PEN):
        if ppr_dict:
            print('загрузка {0}'.format(PEN))
        else:
            print('загрузка {0}, используем название столбцов как образец'.format(PEN))
        
        pen_dict = makeDict(PEN)
    else:
        pen_dict = {}

    if not pen_dict and not ppr_dict:
        ExitWMessage('отсутствуют файлы ПЭН и ППР')

    # создаем книгу для записи ненайденных ИД
    if not os.path.isdir(DIFF_PATH):
        os.mkdir(DIFF_PATH)

    notFoundId = fileExcel(DIFF_PATH + 'no_id_kur_in_ppr_pen.xlsx')
    notFoundId.create('not found')

    # получаем список файлов кураторов и обрабатываем каждый файл
    for fl in listFiles(BASE_DIR):
        print('обработка {0}'.format(fl))

        unique = Unique()

        # если файл с макросами, то окрываем с заданными параметрами
        if fl.endswith('.xlsm'):
            wb = load_workbook(fl, read_only = False, keep_vba = True, data_only = True)
        else:
            wb = load_workbook(fl, data_only = True)

        # выбираем активный лист
        sh = wb.active

        # нужно пропустить строки до Идентификатора
        enable_work = False
        need_len = 0

        # проходим циклом по всем строкам в текущем файле куратора
        for row in sh.iter_rows():

            if enable_work:
                key = row[0].value      # запоминаем текущий идентификатор

                try:
                    unique.add(key)
                except NotUnique as e:
                    ExitWMessage("ошибка в файле {0}\n{1}".format(fl, e.message))

                if len(row) < need_len:
                    row.extend(['' for i in range(need_len - len(row))])

                if key in ppr_dict:
                    # текущий идентификатор есть в ппр
                    notFoundInPpr = False

                    # проходи циклом по массиву индексов заменяемых ячееек
                    for index in PPR_INDEX_I:
                        row[index - 1].value = ppr_dict[key][index - 1]
                else:
                    notFoundInPpr = True

                if key in pen_dict:
                    # есть идентификатор в ппр
                    notFoundInPen = False

                    for index in PEN_INDEX_I:
                        row[index - 1].value = pen_dict[key][index - 1]
                else:
                    notFoundInPen = True

                if notFoundInPen and notFoundInPpr:
                    notFoundId.append([key, fl.replace(BASE_DIR, '')])

                continue

            if row[0].value == 'Идентификатор':
                need_len = len(row)
                enable_work = True

                try:
                    checkHeader([cell.value for cell in row])
                except WrongHeader as e:
                    ExitWMessage("ошибка в файле {0}\n{1}".format(fl, e.message))


        out_filename = makeOut(fl)

        del(unique)

        sh['Y1'] = currDateTime()
        print("сохранение {0}".format(out_filename))
        wb.save(out_filename)

    notFoundId.save()
    del notFoundId
    del pen_dict
    del ppr_dict

checkHeader.template = []

# первый этап: перенос данных из ППР/ПЭН в файлы кураторов
updateKurators()

# второй этап: перенос данных из файлов кураторов в ППР и ПЭН
updatePprPen()

# создание резервной копии с датой
makeBackup()

ExitWMessage("Обработка файлов завершена\n")

