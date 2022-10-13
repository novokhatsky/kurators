# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
import os


IN = 'd:\\tmp\\rubcov\\12'
DELIMENTER = '~'
NAMESHEET = 'Лист 1'


def getvalue(val):
    
    if not val:
        return ''

    return ('{0}'.format(val)).replace('\r', '').replace('\n', '').replace('~', '')


def list_files(base_dir):
    # формируем список файлов с заданной маской в директории
    spisok_file = []
    for i in os.listdir(base_dir):
        if os.path.isfile(os.path.join(base_dir, i)):
            if i.endswith('.xlsx') or i.endswith('.xlsm'):
                spisok_file.append(os.path.join(base_dir, i))

    return spisok_file


def convert(src, dest):
    wb = load_workbook(src, read_only=True, data_only=True)
    sh = wb[NAMESHEET]

    i = 0

    with open(dest, "w", encoding='utf-8') as handle:
        for row in sh.iter_rows():

            i += 1
            if i % 500 == 0:
                print('left {0}'.format(i))

            data = [getvalue(cell.value) for cell in row]

            new_str = DELIMENTER.join(data)

            try:
                handle.write(u'{0}\n'.format(new_str))
            except UnicodeEncodeError:
                print(newStr)
                exit()

    wb.close()


def make_out_name(name):
    fn = name.split('.')
    fn[-1] = 'csv'
    return '.'.join(fn)


files_in = list_files(IN)

for fl in files_in:
    convert(fl, make_out_name(fl))
