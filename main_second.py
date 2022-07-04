# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
import os
import logging
from copy import copy

#BASE_DIR = "d:\\work\\in\\"
#BASE_OUT = "d:\\work\\out\\"
BASE_DIR = "d:\\tmp\\rubcov\\pro\\in\\"
BASE_OUT = "d:\\tmp\\rubcov\\pro\\out\\"

ALL = BASE_OUT + 'out_all.xlsx'

PPR = BASE_DIR + "пэн-ппр\\ППР.xlsx"
PEN = BASE_DIR + "пэн-ппр\\ПЭН.xlsx"

PPR_OUT = BASE_OUT + "пэн-ппр\\ППР.xlsx"
PEN_OUT = BASE_OUT + "пэн-ппр\\ПЭН.xlsx"

PPR_INDEX = [20, 21, 74, 75, 76, 77, 78, 79, 80, 81, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]

PEN_INDEX = [20, 21, 74, 75, 76, 77, 78, 79, 80, 81, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]


LOG_FILE = BASE_DIR + "process.log"

logging.basicConfig(format = u'[%(asctime)s] %(message)s', level = logging.INFO, filename = LOG_FILE)


def out_log(mess):
    logging.info(mess)


def makeOut(filename):
    el = filename.split("\\")

    return BASE_OUT + el[-1]

def makeDict(filename):
    wb = load_workbook(filename, read_only = True)
    sh = wb.active
    data = {} 

    # нужно пропустить три строки
    enable_add = False
    for row in sh.iter_rows():

        if row[0].value == 'Идентификатор':
            enable_add = True
            continue
    
        if enable_add:
            data[row[0].value] = [cell.value for cell in row]

    return data


def listFiles(base_dir):
    # формируем список файлов с заданной маской в директории
    spisok_file = []
    for i in os.listdir(base_dir):
        if os.path.isfile(os.path.join(base_dir, i)):
           if i.endswith('.xlsx'):
                spisok_file.append(os.path.join(base_dir, i))

    return spisok_file


def makeFileKurator(fio):
    fam, nam, sec = fio.split(' ')

    return BASE_OUT + fam + '.xlsx'

    
print('load all')
out_log("загрузка all")

all_dict = makeDict(ALL)

print('load ppr')
wb = load_workbook(PPR)
sh = wb.active

print('seek ppr')
# нужно пропустить три строки
enable_view = False
for row in sh.iter_rows():

    if row[0].value == 'Идентификатор':
        enable_view = True
        continue

    if enable_view:
        key = row[0].value

        if key in all_dict:
            try:
                for index in PPR_INDEX:
                    row[index - 1].value = all_dict[key][index - 1]
            except IndexError:
                continue

print('save ppr')
wb.save(PPR_OUT)
    
print('load pen')
wb = load_workbook(PEN)
sh = wb.active

print('seek pen')
# нужно пропустить три строки
enable_view = False
for row in sh.iter_rows():

    if row[0].value == 'Идентификатор':
        enable_view = True
        continue

    if enable_view:
        key = row[0].value

        if key in all_dict:
            try:
                for index in PEN_INDEX:
                    row[index - 1].value = all_dict[key][index - 1]
            except IndexError:
                continue

print('save pen')
wb.save(PEN_OUT)
    
